"""Tests for pdf-table-stripper — all run without pdfplumber or a real PDF."""

import csv
import json
import os
import tempfile

import pytest

from pdfstripper.models import Table, TableRow
from pdfstripper.extractor import (
    MockPDFExtractor,
    extract_tables,
    _clean_cell,
    _clean_row,
    _is_numeric,
    _is_header_row,
    _drop_empty_rows,
)
from pdfstripper.writer import (
    table_to_csv_string,
    tables_to_csv,
    tables_to_json,
    tables_to_excel,
    _safe_filename,
)

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Model tests
# ---------------------------------------------------------------------------

class TestTable:
    def _sample(self) -> Table:
        return Table(
            page_number=1,
            table_index=0,
            headers=["Date", "Amount", "Description"],
            rows=[
                ["2026-04-01", "1000.00", "Salary"],
                ["2026-04-05", "250.00",  "Rent"],
                ["2026-04-10", "50.00",   "Groceries"],
            ],
            source_file="statement.pdf",
        )

    def test_row_count(self):
        t = self._sample()
        assert t.row_count == 3

    def test_col_count_from_headers(self):
        t = self._sample()
        assert t.col_count == 3

    def test_col_count_from_rows_when_no_headers(self):
        t = Table(page_number=1, table_index=0, rows=[["a", "b", "c"]])
        assert t.col_count == 3

    def test_is_empty_false(self):
        assert not self._sample().is_empty

    def test_is_empty_true(self):
        t = Table(page_number=1, table_index=0)
        assert t.is_empty

    def test_to_dict_rows(self):
        t = self._sample()
        dicts = t.to_dict_rows()
        assert len(dicts) == 3
        assert dicts[0]["Date"] == "2026-04-01"
        assert dicts[0]["Amount"] == "1000.00"

    def test_to_dict_rows_no_headers(self):
        t = Table(page_number=1, table_index=0, rows=[["a", "b"]])
        dicts = t.to_dict_rows()
        assert "col_0" in dicts[0]
        assert dicts[0]["col_0"] == "a"

    def test_summary(self):
        t = self._sample()
        s = t.summary()
        assert "Page 1" in s
        assert "3 rows" in s


# ---------------------------------------------------------------------------
# Extractor helper tests
# ---------------------------------------------------------------------------

class TestCleaners:
    def test_clean_cell_none(self):
        assert _clean_cell(None) == ""

    def test_clean_cell_strips_whitespace(self):
        assert _clean_cell("  hello  ") == "hello"

    def test_clean_cell_collapses_spaces(self):
        assert _clean_cell("hello   world") == "hello world"

    def test_clean_row(self):
        row = [None, "  value  ", "123"]
        assert _clean_row(row) == ["", "value", "123"]

    def test_is_numeric_integer(self):
        assert _is_numeric("1234") is True

    def test_is_numeric_float(self):
        assert _is_numeric("1,234.56") is True

    def test_is_numeric_currency(self):
        assert _is_numeric("$1,234.56") is True

    def test_is_numeric_false(self):
        assert _is_numeric("Description") is False

    def test_is_header_row_true(self):
        assert _is_header_row(["Date", "Description", "Amount"]) is True

    def test_is_header_row_false(self):
        assert _is_header_row(["2026-04-01", "1000.00", "500.00"]) is False

    def test_drop_empty_rows(self):
        rows = [["a", "b"], ["", ""], ["c", "d"]]
        result = _drop_empty_rows(rows)
        assert len(result) == 2


# ---------------------------------------------------------------------------
# MockPDFExtractor tests
# ---------------------------------------------------------------------------

class TestMockExtractor:
    def test_returns_tables(self):
        tables = MockPDFExtractor.extract()
        assert len(tables) >= 2

    def test_first_table_has_headers(self):
        tables = MockPDFExtractor.extract()
        assert tables[0].headers == ["Date", "Description", "Debit", "Credit", "Balance"]

    def test_first_table_has_rows(self):
        tables = MockPDFExtractor.extract()
        assert tables[0].row_count > 0

    def test_tables_have_page_numbers(self):
        tables = MockPDFExtractor.extract()
        assert all(t.page_number >= 1 for t in tables)

    def test_extract_tables_mock_flag(self):
        tables = extract_tables("any.pdf", use_mock=True)
        assert len(tables) >= 2

    def test_extract_tables_fallback_without_pdfplumber(self):
        # Without pdfplumber installed, should fall back to mock
        import sys
        # Temporarily hide pdfplumber if installed
        import unittest.mock as mock
        with mock.patch.dict(sys.modules, {"pdfplumber": None}):
            tables = extract_tables("any.pdf")
        assert len(tables) >= 2


# ---------------------------------------------------------------------------
# Writer tests
# ---------------------------------------------------------------------------

class TestWriter:
    def _sample_table(self) -> Table:
        return Table(
            page_number=1,
            table_index=0,
            headers=["Date", "Amount"],
            rows=[["2026-04-01", "1000.00"], ["2026-04-05", "250.00"]],
            source_file="statement.pdf",
        )

    def test_table_to_csv_string_has_headers(self):
        t = self._sample_table()
        csv_str = table_to_csv_string(t)
        assert "Date,Amount" in csv_str

    def test_table_to_csv_string_has_rows(self):
        t = self._sample_table()
        csv_str = table_to_csv_string(t)
        assert "2026-04-01" in csv_str

    def test_tables_to_csv_creates_files(self):
        tables = MockPDFExtractor.extract()
        with tempfile.TemporaryDirectory() as d:
            paths = tables_to_csv(tables, output_dir=d)
            assert len(paths) == len(tables)
            for p in paths:
                assert os.path.exists(p)

    def test_tables_to_csv_single_file(self):
        tables = MockPDFExtractor.extract()
        with tempfile.TemporaryDirectory() as d:
            paths = tables_to_csv(tables, output_dir=d, single_file=True)
            assert len(paths) == 1
            assert os.path.exists(paths[0])

    def test_csv_content_is_valid(self):
        t = self._sample_table()
        with tempfile.TemporaryDirectory() as d:
            paths = tables_to_csv([t], output_dir=d)
            with open(paths[0]) as f:
                rows = list(csv.reader(f))
            assert rows[0] == ["Date", "Amount"]
            assert rows[1][0] == "2026-04-01"

    def test_tables_to_json(self):
        tables = MockPDFExtractor.extract()
        result = json.loads(tables_to_json(tables))
        assert isinstance(result, list)
        assert len(result) == len(tables)
        assert "headers" in result[0]
        assert "rows" in result[0]

    def test_safe_filename(self):
        name = _safe_filename("statement.pdf", 1, 0)
        assert name == "statement_page1_table1.csv"

    def test_safe_filename_special_chars(self):
        name = _safe_filename("my report (2026).pdf", 2, 1)
        assert ".csv" in name
        assert "page2" in name

    def test_tables_to_excel_creates_files(self):
        tables = MockPDFExtractor.extract()
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "out.xlsx")
            tables_to_excel(tables, out)
            assert os.path.exists(out)

    def test_tables_to_excel_requires_dependency(self):
        tables = MockPDFExtractor.extract()

        import sys
        from unittest import mock

        with mock.patch.dict(sys.modules, {"openpyxl": None}):
            with pytest.raises(RuntimeError):
                tables_to_excel(tables, "out.xlsx")

    def test_excel_content(self):
        tables = MockPDFExtractor.extract()

        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "out.xlsx")
            tables_to_excel(tables, out)

            wb = load_workbook(out)
            sheets = wb.sheetnames

            assert len(sheets) == len(tables)

            ws = wb[sheets[0]]
            assert ws.cell(row=1, column=1).value is not None

    def test_excel_headers_are_bold(self):
        tables = MockPDFExtractor.extract()

        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "out.xlsx")
            tables_to_excel(tables, out)

            wb = load_workbook(out)
            ws = wb[wb.sheetnames[0]]

            for col_idx in range(1, len(tables[0].headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                assert cell.font.bold is True

    def test_excel_sheet_name_truncation(self):
        long_table = Table(
            page_number=1,
            table_index=0,
            headers=["a"],
            rows=[],
            source_file="very_long_filename_that_exceeds_normal_limits.pdf"
        )

        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "out.xlsx")
            tables_to_excel([long_table], out)

            wb = load_workbook(out)

            assert len(wb.sheetnames[0]) <= 31